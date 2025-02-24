from tkinter import Label, Button, Entry, Tk, messagebox
from pandas import read_excel, ExcelWriter, isna
from openpyxl import Workbook
from math import floor
from tkcalendar import DateEntry
from datetime import datetime, timedelta
import babel, babel.numbers

screen = Tk()
screen.config(width=400, height=400, padx=200, pady=50, bg='#070F2B')

current_date = datetime.now()
month = current_date.strftime('%B')
file_name = f'{month}_Data.xlsx'


def get_date_today(event, items, item_no):
    global selected_date
    selected_date = cal.get_date()
    day = selected_date.strftime('%d')
    selected_date_str = selected_date.strftime('%d-%m-%Y')

    data2 = read_excel(file_name, sheet_name="Sheet2")
    index_today = data2.index[data2["Date"] == selected_date_str]

    amount_left_input.delete(0, 'end')
    new_item_input.delete(0, 'end')
    used_input.delete(0, 'end')
    left_input.delete(0, 'end')

    if day != '01':

        day_before = selected_date - timedelta(days=1)
        day_before = day_before.strftime("%d-%m-%Y")
        index_old = data2.index[data2['Date'] == day_before]

        last_day_remaining = float(data2.loc[index_old+5, items[item_no]].iloc[0])
        if isna(last_day_remaining):
            last_day_remaining = float(data2.loc[index_today, items[item_no]].iloc[0])
            if isna(last_day_remaining):
                pass
            else:
                amount_left_input.insert(0, last_day_remaining)
        else:
            amount_left_input.insert(0, last_day_remaining)

        new_today = float(data2.loc[index_today+1][items[item_no]].iloc[0])
        if isna(new_today):
            pass
        else:
            new_item_input.insert(0, new_today)

        sale_today = float(data2.loc[index_today + 3][items[item_no]].iloc[0])
        if isna(sale_today):
            pass
        else:
            used_input.insert(0, sale_today)

        actual_today = float(data2.loc[index_today+5][items[item_no]].iloc[0])
        if isna(actual_today):
            pass
        else:
            left_input.insert(0, actual_today)


def get_dates_of_current_month():
    today = datetime.today()
    start_date = today.replace(day=1)
    next_month = start_date.replace(month=start_date.month % 12 + 1, day=1)
    days_in_month = (next_month - start_date).days
    dates = [start_date + timedelta(days=i) for i in range(days_in_month)]
    return dates


def focus_next(event):
    current_index = entry_boxes.index(event.widget)
    if current_index < len(entry_boxes) - 1:
        entry_boxes[current_index + 1].focus()


def destroy_all():
    for widget in screen.winfo_children():
        widget.destroy()


def new_items_add():
    destroy_all()
    name = Label(screen, text="Item name:", bg='#9290C3')
    name.grid(row=0, column=0, pady=10)
    name_input = Entry(screen, width=15, bg='#9290C3')
    name_input.grid(row=0, column=1, pady=10, padx=10)
    name_input.focus_set()

    submit = Button(screen, text='Submit', command=lambda: add_new_excel(name_input), bg='#9290C3')
    submit.grid(row=1, column=1, padx=15, pady=15)

    back = Button(screen, text="Back", command=initiate, bg='#9290C3')
    back.grid(row=1, column=0)


def add_new_excel(name_input):
    data = read_excel(file_name, sheet_name="Sheet1")
    data2 = read_excel(file_name, sheet_name="Sheet2")
    data2[name_input.get().capitalize()] = None
    data[name_input.get().capitalize()] = None
    with ExcelWriter(file_name, engine='openpyxl', mode='a') as writer:
        writer.book.remove(writer.book["Sheet1"])
        writer.book.remove(writer.book["Sheet2"])
        data.to_excel(writer, sheet_name="Sheet1", index=False)
        data2.to_excel(writer, sheet_name="Sheet2", index=False)
    initiate()


def entry(item_no, items):
    try:
        old = float(amount_left_input.get())
        new = float(new_item_input.get())
        used = float(used_input.get())
        left = float(left_input.get())
        selected_date_str = selected_date.strftime('%d-%m-%Y')
    except ValueError:
        messagebox.showwarning("Error", "Please Enter a Number")
    except NameError:
        messagebox.showwarning("Error", "Please Enter a Date")
    else:
        loss = -(old + new - used - left)
        data = read_excel(file_name, sheet_name="Sheet1")
        data2 = read_excel(file_name, sheet_name="Sheet2")
        indices2 = data2.index[data2["Date"] == selected_date_str]
        indices3 = data2.index[data2['Date'] == (selected_date + timedelta(days=1)).strftime('%d-%m-%Y')]

        data2.loc[indices2, items[item_no]] = old
        data2.loc[indices2+1, items[item_no]] = new
        data2.loc[indices2+2, items[item_no]] = old+new
        data2.loc[indices2+3, items[item_no]] = used
        data2.loc[indices2+4, items[item_no]] = old+new-used
        data2.loc[indices2+5, items[item_no]] = left
        data2.loc[indices2+6, items[item_no]] = loss

        indices = data.index[data['Date'] == selected_date_str]
        data.loc[indices, items[item_no]] = loss

        if (selected_date + timedelta(days=1)).strftime('%d') == '01':
            pass
        else:
            data2.loc[indices3, items[item_no]] = float(data2.loc[indices2+5, items[item_no]].iloc[0])

        try:
            new_1 = float(data2.loc[indices3+1, items[item_no]].iloc[0])
            new_2 = float(data2.loc[indices3+3, items[item_no]].iloc[0])
            new_3 = float(data2.loc[indices3+5, items[item_no]].iloc[0])
        except ValueError:
            pass
        else:
            data2.loc[indices3+2, items[item_no]] = float(data2.loc[indices3+1, items[item_no]].iloc[0]) + float(data2.loc[indices3, items[item_no]].iloc[0])
            data2.loc[indices3 + 4, items[item_no]] = float(data2.loc[indices3 + 2, items[item_no]].iloc[0]) - float(data2.loc[indices3+3, items[item_no]].iloc[0])
            data2.loc[indices3 + 6, items[item_no]] = float(data2.loc[indices3 + 5, items[item_no]].iloc[0]) - float(data2.loc[indices3 + 4, items[item_no]].iloc[0])
            temp = data.index[data['Date'] == (selected_date + timedelta(days=1)).strftime('%d-%m-%Y')]
            data.loc[temp, items[item_no]] = data2.loc[indices3 + 6, items[item_no]]

        with ExcelWriter(file_name, engine='openpyxl', mode='a') as writer:
            writer.book.remove(writer.book["Sheet1"])
            writer.book.remove(writer.book["Sheet2"])
            data.to_excel(writer, sheet_name="Sheet1", index=False)
            data2.to_excel(writer, sheet_name="Sheet2", index=False)

        initiate()


def crazy(item_no, items):
    destroy_all()
    global amount_left_input, new_item_input, used_input, left_input, entry_boxes, cal

    amount_left = Label(screen, text="Amount Left:", bg='#9290C3')
    amount_left.grid(row=1, column=0, pady=10)
    amount_left_input = Entry(screen, width=15, bg='#9290C3')
    amount_left_input.grid(row=1, column=1, pady=10, padx=10)
    amount_left_input.focus_set()
    new_items = Label(screen, text='New Items:', bg='#9290C3')
    new_items.grid(row=2, column=0, pady=10)
    new_item_input = Entry(screen, width=15, bg='#9290C3')
    new_item_input.grid(row=2, column=1, pady=10, padx=10)
    used = Label(screen, text='Amount Used:', bg='#9290C3')
    used.grid(row=3, column=0, pady=10)
    used_input = Entry(screen, width=15, bg='#9290C3')
    used_input.grid(row=3, column=1, pady=10, padx=10)
    left = Label(screen, text='Amount Left:', bg='#9290C3')
    left.grid(row=4, column=0, pady=10)
    left_input = Entry(screen, width=15, bg='#9290C3')
    left_input.grid(row=4, column=1, pady=10, padx=10)

    cal = DateEntry(screen, date_pattern='dd-mm-yyyy', padx=10, pady=10)
    cal.grid(row=0, column=1)
    cal.bind("<<DateEntrySelected>>", lambda event, items=items, item_no=item_no: get_date_today(event, items, item_no))

    entry_boxes = [amount_left_input, new_item_input, used_input, left_input]

    submit = Button(screen, text='Submit', command=lambda: entry(item_no, items), bg='#9290C3')
    submit.grid(row=5, column=1, padx=15, pady=15)

    back = Button(screen, text="Back", command=initiate, bg='#9290C3')
    back.grid(row=5, column=0)

    for entry_1 in entry_boxes:
        entry_1.bind("<Return>", focus_next)


def initiate():
    destroy_all()
    try:
        data = read_excel(file_name, sheet_name="Sheet1")
    except FileNotFoundError:
        wb = Workbook()
        ws = wb.active
        ws.title = 'Sheet1'
        ws.cell(row=1, column=1, value='Date')
        ws2 = wb.create_sheet(title="Sheet2")
        ws2.cell(row=1, column=1, value='Date')
        dates = get_dates_of_current_month()
        for i, date in enumerate(dates, start=2):  # Start from row 2 to account for the header
            ws.cell(row=i, column=1, value=date.strftime('%d-%m-%Y'))
            if i == 2:
                ws2.cell(row=2, column=1, value=date.strftime('%d-%m-%Y'))
            else:
                ws2.cell(row=(i-2)*7+2, column=1, value=date.strftime('%d-%m-%Y'))
            for j in range(0, len(titels)):
                ws2.cell(row=(i-2)*7+2+j, column=2, value=titels[j])
        wb.save(file_name)
        data = read_excel(file_name, sheet_name="Sheet1")

    items = list(data.columns)
    buttons = []
    column = 0
    row = 0
    items.remove('Date')

    for i in range(0, len(items)):
        if row == 4:
            column += 1
            row = 0
        buttons.append(Button(screen, text=items[i], command=lambda i=i: crazy(i, items), width=20, height=3, bg='#9290C3'))
        buttons[i].grid(row=row, column=column, padx=10, pady=10)
        row += 1

    new_items = Button(screen, text="New Items", command=new_items_add, width=20, height=3, bg='#535C91')
    new_items.grid(row=5, column=floor(column/2), padx=10, pady=20)


titels = ["Leftover", "New", "Total", "Sale", "Remaining", "Actual Remaining", "Loss/Profit"]
initiate()

screen.mainloop()
